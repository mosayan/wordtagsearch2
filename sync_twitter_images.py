import requests
import os
import sys
import datetime
import re
import time
from openpyxl import load_workbook
from dotenv import load_dotenv

# .envの読み込み
load_dotenv()

# ==========================================
# 設定の取得とチェック
# ==========================================
BEARER_TOKEN = os.getenv('BEARER_TOKEN')
SAVE_DIR = os.getenv('SAVE_DIR')
EXCEL_FILE = os.getenv('EXCEL_FILE')
SEARCH_LIST_SHEET = os.getenv('SEARCH_LIST_SHEET')

# C列が空欄だった場合の「デフォルト値」として .env の値を使用
try:
    DEFAULT_MAX_RESULTS = int(os.getenv('MAX_RESULTS', 10))
except ValueError:
    print("エラー: .envの MAX_RESULTS は数値で指定してください。")
    sys.exit(1)

if not all([BEARER_TOKEN, SAVE_DIR, EXCEL_FILE, SEARCH_LIST_SHEET]):
    print("エラー: .envファイルに必要な設定が不足しています。")
    sys.exit(1)

if not (10 <= DEFAULT_MAX_RESULTS <= 100):
    print("エラー: .envの MAX_RESULTS は 10 から 100 の間で指定してください。")
    sys.exit(1)

if not os.path.exists(SAVE_DIR):
    os.makedirs(SAVE_DIR)

def format_tag_name(query):
    name = re.sub(r'[\\/?*\[\]:|<>"]', '_', query)
    return name.strip()[:31]

def get_search_queries():
    if not os.path.exists(EXCEL_FILE):
        print(f"エラー: 指定されたExcelファイルが見つかりません [{EXCEL_FILE}]")
        sys.exit(1)
    
    wb = load_workbook(EXCEL_FILE, data_only=True)
    if SEARCH_LIST_SHEET not in wb.sheetnames:
        print(f"エラー: 検索リストシートが見つかりません [{SEARCH_LIST_SHEET}]")
        sys.exit(1)
        
    ws = wb[SEARCH_LIST_SHEET]
    queries = []
    
    # 2行目から読み込み
    for row in range(2, ws.max_row + 1):
        tag_val = ws.cell(row=row, column=2).value      # B列: 検索タグ
        max_res_val = ws.cell(row=row, column=3).value  # C列: max_results
        
        if tag_val:
            try:
                # C列が空欄でない場合はその数値を、空欄ならデフォルト値を使用
                if max_res_val is not None:
                    current_max = int(max_res_val)
                else:
                    current_max = DEFAULT_MAX_RESULTS
                
                # APIの仕様（10〜100）に強制的に丸める
                current_max = max(10, min(100, current_max))
            except (ValueError, TypeError):
                # max_resultsに文字などが入っていたらデフォルト値にする
                current_max = DEFAULT_MAX_RESULTS

            # 連想配列リストに追加
            queries.append({
                'query': str(tag_val),
                'max_results': current_max
            })
            
    return queries

# 引数 query_info には辞書型 {'query': '#タグ', 'max_results': 50} が入る
def search_recent_tweets_with_images(query_info, since_id=None):
    # 呼び出すAPIのエンドポイント、およびパラメータをセット
    url = "https://api.twitter.com/2/tweets/search/recent"
    headers = {"Authorization": f"Bearer {BEARER_TOKEN}"}
    params = {
        'query': f"{query_info['query']} has:media -is:retweet",
        'expansions': 'attachments.media_keys',
        'media.fields': 'url',
        'max_results': query_info['max_results']
    }
    
    if since_id:
        params['since_id'] = since_id

    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        print(f"API Error for {query_info['query']}: {response.status_code} {response.text}")
        return None
    return response.json()

def download_image(url, tweet_id, index, save_dir):
    # URLの「?」以降（既存のサイズ指定パラメータなど）を一旦切り捨てる
    base_url = url.split('?')[0]
    
    # 拡張子を抽出してファイル名を作成
    ext = base_url.split('.')[-1]
    filename = f"{tweet_id}_{index}.{ext}"
    filepath = os.path.join(save_dir, filename)
    
    # ファイルのオリジナルサイズ指定するname=origのパラメータを付与
    high_res_url = f"{base_url}?name=orig"
    
    # 高画質URLに対してダウンロードを実行
    response = requests.get(high_res_url, stream=True)
    if response.status_code == 200:
        with open(filepath, 'wb') as f:
            for chunk in response.iter_content(1024):
                f.write(chunk)
        return filename
    return None

def main():
    # 検索クエリのリストを取得
    queries = get_search_queries()
    
    if not queries:
        print("検索クエリがありません。シートのB列にタグを入力してください。")
        sys.exit(0)

    wb = load_workbook(EXCEL_FILE)
    today_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for index, q_info in enumerate(queries):
        query = q_info['query']
        current_max = q_info['max_results']
        tag_name = format_tag_name(query)
        
        print(f"\n--- 検索開始: {query} (取得件数: {current_max}件 / 保存名: {tag_name}) ---")
        
        tag_dir = os.path.join(SAVE_DIR, tag_name)
        os.makedirs(tag_dir, exist_ok=True)
        
        if tag_name in wb.sheetnames:
            ws = wb[tag_name]
        else:
            # 新しいタグ用のシートを追加
            ws = wb.create_sheet(title=tag_name)
            ws.append(['Tweet ID', 'Image URL', 'Filename', 'Date'])
        
        # 既存のTweet ID一覧を取得(ヘッダー(Tweet ID)を除く)
        # 集合内包表記
        existing_ids = {str(cell.value) for cell in ws['A'] if cell.value != 'Tweet ID'}
        
        if existing_ids:
            # すでに保持しているIDから一番最新のIDを取得
            latest_id = str(max([int(tid) for tid in existing_ids]))
            print(f"  -> 前回取得した最新ID ({latest_id}) 以降の差分を検索します。")
        else:
            latest_id = None
        
        api_response = search_recent_tweets_with_images(q_info, latest_id)
        
        if not api_response or 'data' not in api_response or 'includes' not in api_response:
            print("  -> 新しい画像付きの投稿は見つかりませんでした。")
        else:
            media_dict = {}
            for media in api_response['includes'].get('media', []):
                if media['type'] == 'photo':
                    media_dict[media['media_key']] = media['url']

            download_count = 0

            for tweet in api_response['data']:
                tweet_id = tweet['id']
                if tweet_id in existing_ids:
                    continue
                    
                if 'attachments' in tweet and 'media_keys' in tweet['attachments']:
                    for img_index, media_key in enumerate(tweet['attachments']['media_keys']):
                        if media_key in media_dict:
                            image_url = media_dict[media_key]
                            filename = download_image(image_url, tweet_id, img_index, tag_dir)
                            
                            if filename:
                                ws.append([tweet_id, image_url, filename, today_str])
                                print(f"  ✅ 保存完了: {filename}")
                                download_count += 1
                                existing_ids.add(tweet_id)
            
            print(f"  -> {download_count} 件の新しい画像を保存しました。")

        print(f"  [中間セーブ] '{tag_name}' の結果を保存中...")
        while True:
            try:
                wb.save(EXCEL_FILE)
                break # 保存成功
            except PermissionError:
                print(f"\nエラー: Excelファイル ({EXCEL_FILE}) が開かれたままです。")
                input("   Excelを閉じてから [Enter] キーを押して保存を再試行してください...")

        if index < len(queries) - 1:
            print("  [API制限対策] 次の検索まで3秒待機します...")
            time.sleep(3)

    print("\nデータの保存処理を行っています...")
    while True:
        try:
            wb.save(EXCEL_FILE)
            break
        except PermissionError:
            print(f"\nエラー: Excelファイル ({EXCEL_FILE}) が開かれたままになっています。")
            input("   Excelを閉じてから、この画面で [Enter] キーを押して再試行してください...")

    print("\nすべての処理が完了しました。")

if __name__ == "__main__":
    main()